"""MCP server for network drive file access.

Provides 6 tools for reading/writing files on the shared network drive used by
the accounting team (Windows N:\\ or Mac /Volumes/Network equivalent).

Environment variables
---------------------
FS_DRIVE_ROOT   Root path of the network drive (default: N:\\ on Windows,
                /Volumes/Network on Mac).  The server never escapes this root.
"""

from __future__ import annotations

import csv
import io
import os
import sys
from pathlib import Path

from mcp.server.fastmcp import FastMCP

mcp = FastMCP("sentient-ledger-filesystem")

# ---------------------------------------------------------------------------
# Drive root resolution
# ---------------------------------------------------------------------------

_WINDOWS_DEFAULT = r"N:\\"
_MAC_DEFAULT = "/Volumes/Network"


def _drive_root() -> Path:
    """Return the configured network drive root."""
    env = os.environ.get("FS_DRIVE_ROOT", "")
    if env:
        return Path(env)
    if sys.platform == "win32":
        return Path(_WINDOWS_DEFAULT)
    return Path(_MAC_DEFAULT)


def _safe_path(relative: str) -> Path:
    """Resolve a relative path under drive root; raise ValueError if it escapes."""
    root = _drive_root()
    target = (root / relative).resolve()
    if not str(target).startswith(str(root.resolve())):
        raise ValueError(f"Path {relative!r} escapes drive root {root}")
    return target


# ---------------------------------------------------------------------------
# Tools
# ---------------------------------------------------------------------------


@mcp.tool()
def list_files(relative_dir: str = "", pattern: str = "*") -> dict:
    """List files in a directory on the network drive.

    Args:
        relative_dir: Directory path relative to the drive root.  Empty string
                      means the root itself.
        pattern:      Glob pattern to filter filenames (default ``*``).

    Returns:
        Dictionary with ``files`` (list of relative paths) and ``count``.
    """
    try:
        directory = _safe_path(relative_dir)
        if not directory.exists():
            return {"files": [], "count": 0, "error": f"Directory not found: {relative_dir}"}
        matches = [
            str(p.relative_to(_drive_root()))
            for p in sorted(directory.glob(pattern))
            if p.is_file()
        ]
        return {"files": matches, "count": len(matches)}
    except ValueError as exc:
        return {"files": [], "count": 0, "error": str(exc)}
    except Exception as exc:
        return {"files": [], "count": 0, "error": f"Unexpected error: {exc}"}


@mcp.tool()
def read_csv(relative_path: str, max_rows: int = 1000) -> dict:
    """Read a CSV file from the network drive.

    Args:
        relative_path: File path relative to the drive root.
        max_rows:      Maximum number of data rows to return (default 1000).

    Returns:
        Dictionary with ``headers`` (list), ``rows`` (list of dicts), and
        ``total_rows`` count.
    """
    try:
        path = _safe_path(relative_path)
        if not path.exists():
            return {"headers": [], "rows": [], "total_rows": 0, "error": "File not found"}
        content = path.read_text(encoding="utf-8-sig")
        reader = csv.DictReader(io.StringIO(content))
        rows = []
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            rows.append(dict(row))
        return {
            "headers": reader.fieldnames or [],
            "rows": rows,
            "total_rows": len(rows),
        }
    except ValueError as exc:
        return {"headers": [], "rows": [], "total_rows": 0, "error": str(exc)}
    except Exception as exc:
        return {"headers": [], "rows": [], "total_rows": 0, "error": f"Unexpected error: {exc}"}


@mcp.tool()
def read_excel(relative_path: str, sheet_name: str = "", max_rows: int = 1000) -> dict:
    """Read an Excel file from the network drive.

    Requires ``openpyxl`` (already a project dependency).

    Args:
        relative_path: File path relative to the drive root.
        sheet_name:    Sheet to read; defaults to the first sheet.
        max_rows:      Maximum data rows to return (default 1000).

    Returns:
        Dictionary with ``headers``, ``rows``, ``sheet``, and ``total_rows``.
    """
    try:
        import openpyxl

        path = _safe_path(relative_path)
        if not path.exists():
            return {"headers": [], "rows": [], "total_rows": 0, "error": "File not found"}

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        rows_iter = ws.iter_rows(values_only=True)

        header_row = next(rows_iter, None)
        if header_row is None:
            return {"headers": [], "rows": [], "sheet": ws.title, "total_rows": 0}

        headers = [str(h) if h is not None else "" for h in header_row]
        rows = []
        for i, row in enumerate(rows_iter):
            if i >= max_rows:
                break
            rows.append(dict(zip(headers, [str(v) if v is not None else "" for v in row])))

        return {"headers": headers, "rows": rows, "sheet": ws.title, "total_rows": len(rows)}
    except ValueError as exc:
        return {"headers": [], "rows": [], "total_rows": 0, "error": str(exc)}
    except Exception as exc:
        return {"headers": [], "rows": [], "total_rows": 0, "error": f"Unexpected error: {exc}"}


@mcp.tool()
def write_csv(relative_path: str, headers: list[str], rows: list[dict]) -> dict:
    """Write rows to a CSV file on the network drive.

    Creates parent directories as needed.  Overwrites the file if it exists.

    Args:
        relative_path: Destination path relative to the drive root.
        headers:       Column names in desired order.
        rows:          List of dicts; extra keys are ignored.

    Returns:
        Dictionary with ``written_rows`` count and ``path``.
    """
    try:
        path = _safe_path(relative_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.DictWriter(fh, fieldnames=headers, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
        return {"written_rows": len(rows), "path": str(path)}
    except ValueError as exc:
        return {"written_rows": 0, "error": str(exc)}
    except Exception as exc:
        return {"written_rows": 0, "error": f"Unexpected error: {exc}"}


@mcp.tool()
def file_exists(relative_path: str) -> dict:
    """Check whether a file exists on the network drive.

    Args:
        relative_path: File path relative to the drive root.

    Returns:
        Dictionary with ``exists`` (bool) and ``size_bytes`` (int, 0 if missing).
    """
    try:
        path = _safe_path(relative_path)
        exists = path.exists() and path.is_file()
        return {"exists": exists, "size_bytes": path.stat().st_size if exists else 0}
    except ValueError as exc:
        return {"exists": False, "size_bytes": 0, "error": str(exc)}
    except Exception as exc:
        return {"exists": False, "size_bytes": 0, "error": f"Unexpected error: {exc}"}


@mcp.tool()
def get_latest_file(relative_dir: str, pattern: str = "*") -> dict:
    """Return the most recently modified file matching a pattern.

    Args:
        relative_dir: Directory path relative to the drive root.
        pattern:      Glob pattern (default ``*``).

    Returns:
        Dictionary with ``path``, ``modified_at`` (ISO-8601), and ``size_bytes``.
        ``path`` is empty string if no files matched.
    """
    try:
        directory = _safe_path(relative_dir)
        if not directory.exists():
            return {"path": "", "modified_at": "", "size_bytes": 0, "error": "Directory not found"}

        candidates = [p for p in directory.glob(pattern) if p.is_file()]
        if not candidates:
            return {"path": "", "modified_at": "", "size_bytes": 0}

        latest = max(candidates, key=lambda p: p.stat().st_mtime)
        import datetime as dt

        mtime = dt.datetime.fromtimestamp(latest.stat().st_mtime, tz=dt.timezone.utc)
        return {
            "path": str(latest.relative_to(_drive_root())),
            "modified_at": mtime.isoformat(),
            "size_bytes": latest.stat().st_size,
        }
    except ValueError as exc:
        return {"path": "", "modified_at": "", "size_bytes": 0, "error": str(exc)}
    except Exception as exc:
        return {"path": "", "modified_at": "", "size_bytes": 0, "error": f"Unexpected error: {exc}"}


if __name__ == "__main__":
    mcp.run()
